"""Parse OMB Table 4.1 into a curated agency spending CSV.

Reads: data/govt_spending/raw/hist04z1.xlsx
Writes: data/govt_spending/us_agency.csv

The Excel layout (as confirmed by inspection):
- Row 1: Title
- Row 2: Units note
- Row 3: Header row — column 1 is "Department or other unit", remaining columns are
          fiscal years as plain integers (1962, 1963, ..., 2025) or strings like
          "2026 estimate"
- Rows 4-39: Agency data, all at indent=0 except rows 37-38 (sub-lines of
             Undistributed Offsetting Receipts, indent=1) and row 39 (Total, indent=3)
- Values are floats (millions USD) or '...........' strings for missing/NA data

Agency name conventions:
- "Social Security Administration (On-Budget)" and "(Off-Budget)" are split rows;
  we keep Off-Budget (the main benefit program) and drop On-Budget.
- "Other Independent Agencies (On-Budget)" and "(Off-Budget)" similarly; we keep
  Off-Budget and drop On-Budget.
- "Undistributed Offsetting Receipts" and its sub-rows are excluded (negative
  offsets, not agency spending).
- "Total outlays" and "Allowances" are excluded.
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

TARGET_YEARS = [2015, 2020, 2023, 2025]

RAW_PATH = Path("data/govt_spending/raw/hist04z1.xlsx")
OUTPUT_DIR = Path("data/govt_spending")

# Substrings that identify rows to skip entirely (case-insensitive)
SKIP_SUBSTRINGS = [
    "total outlays",
    "undistributed",
    "allowances",
    "(on-budget)",          # skip on-budget split rows; keep off-budget
]

# Rename off-budget agencies to clean names (drop the suffix)
RENAME_MAP = {
    "Social Security Administration (Off-Budget)": "Social Security Administration",
    "Other Independent Agencies (Off-Budget)": "Other Independent Agencies",
}


def parse_omb_agency() -> Path:
    """Parse OMB Table 4.1 and write agency spending CSV."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(RAW_PATH, data_only=True)
    ws = wb.active

    # --- Locate header row and fiscal year columns ---
    header_row = None
    year_cols: dict[int, int] = {}   # fiscal_year -> 1-based column index

    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value is None:
                continue
            val_str = str(cell.value).strip()
            # Headers are plain ints ("2015") or "YYYY estimate"
            for fy in TARGET_YEARS:
                if val_str == str(fy) or val_str.startswith(str(fy) + " "):
                    year_cols[fy] = cell.column
                    header_row = cell.row

    if not year_cols:
        raise ValueError(
            "Could not find fiscal year columns in OMB Excel file. "
            "The file layout may have changed — inspect manually."
        )

    print(f"Header row: {header_row}, fiscal year columns: {year_cols}")

    # --- Extract agency data ---
    # agencies[fy] is a list of (agency_name, amount_millions)
    agencies: dict[int, list[tuple[str, float]]] = {fy: [] for fy in TARGET_YEARS}

    for row in ws.iter_rows(min_row=header_row + 1):
        name_cell = row[0]   # column A (0-indexed)

        # Skip blank rows
        if not name_cell.value or not isinstance(name_cell.value, str):
            continue

        name = name_cell.value.strip()
        name_lower = name.lower()

        # Skip indented sub-rows (Undistributed sub-lines, Total)
        indent = name_cell.alignment.indent if name_cell.alignment else 0
        if indent and indent > 0:
            continue

        # Skip by substring match
        if any(pat in name_lower for pat in SKIP_SUBSTRINGS):
            continue

        # Apply clean name if needed
        display_name = RENAME_MAP.get(name, name)

        for fy, col in year_cols.items():
            val = row[col - 1].value   # convert 1-based col to 0-based index
            # Skip missing/placeholder values (OMB uses '...........' strings)
            if val is None or not isinstance(val, (int, float)):
                continue
            agencies[fy].append((display_name, float(val)))

    # --- Sort, compute top 10 and percentages, write CSV ---
    out_path = OUTPUT_DIR / "us_agency.csv"
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "fiscal_year", "rank", "agency",
            "amount_millions", "amount_billions", "pct_of_total",
        ])

        for fy in TARGET_YEARS:
            rows = agencies[fy]
            if not rows:
                print(f"Warning: no data found for FY{fy}")
                continue

            # Sort by amount descending; negative values (receipts) go to the bottom
            sorted_rows = sorted(rows, key=lambda x: -x[1])
            # Total for percentage: sum of positive outlays only
            total = sum(v for _, v in rows if v > 0)

            for rank, (name, amount) in enumerate(sorted_rows[:10], 1):
                writer.writerow([
                    fy,
                    rank,
                    name,
                    round(amount),
                    round(amount / 1000, 1),
                    round(amount / total * 100, 1) if total else 0,
                ])
            print(f"FY{fy}: top agency = {sorted_rows[0][0]} ({sorted_rows[0][1]/1000:.0f}B), "
                  f"total positive outlays = {total/1000:.0f}B")

    print(f"\nWrote {out_path}")
    wb.close()
    return out_path


if __name__ == "__main__":
    parse_omb_agency()
